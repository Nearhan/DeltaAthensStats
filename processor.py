from openpyxl.workbook import Workbook
import os, datetime, re, openpyxl


#############
## settings ##
#############

settings = { 'current_directory' : os.path.join(os.getcwd()+'/reports', datetime.date.today().isoformat()),
'date' : datetime.date.strftime(datetime.date.today(), '%D').replace('/', '-'),

}


class Processor(object):
	global settings


	def __init__(self):
		self.configure(settings)


	def configure(self, settings):

		for key, value in settings.iteritems():
			setattr(self, key, value)


	def make_call_to_apple_script(self):
		pass


	def gather_all_files_and_split(self):
		self.all_xlsx = []
		self.business_xlsx = []
		self.regular_xlsx= []

		all_files = os.listdir(self.current_directory)
		for single_file in all_files:
			if re.search(r'.xlsx$', single_file):
				dir_file = os.path.join(self.current_directory, single_file)
				self.all_xlsx.append(dir_file)
		for xlsx in self.all_xlsx:
			if re.search(r'Bus', xlsx):
				self.business_xlsx.append(xlsx)
			else:
				self.regular_xlsx.append(xlsx)


	def sort_all_files(self):
		self.sort_business_xlsx()
		self.sort_regular_xlsx()


	def sort_business_xlsx(self):
		patterns = [r'NEW', r'Reopened', r'As', r'FIXED', r'CANCELLED', r'CLOSED', r'UNR']
		self.business_xlsx = self.sort_files(patterns=patterns, arr=self.business_xlsx)


	def sort_regular_xlsx(self):
		patterns = [r'Open\+bugs', r'Resolved\+I', r'Closed\+I', r'Open\+QC', r'Resolved\+QC', r'Closed\+QC']
		self.regular_xlsx = self.sort_files(patterns=patterns, arr=self.regular_xlsx)


	def sort_files(self, patterns=[], arr=[]):
		sorted_files = []
		while patterns:
			pat = patterns.pop()
			for xlsx in arr:
				if re.search(pat, xlsx):
					sorted_files.append(xlsx)

		#reverse array
		arr = sorted_files[::-1]
		return arr


	def create_both_workbooks(self):
		self.create_business_workbook()
		self.create_regular_workbook()


	def create_business_workbook(self):

		workbook = Workbook()

		for xlsx in self.business_xlsx:
			wb = openpyxl.load_workbook(xlsx)
			ws = wb.get_active_sheet()

			if re.search(r'NEW', xlsx):
				ws.title = 'New'
			elif re.search(r'Reopened', xlsx):
				ws.title = 'Reopened'
			elif re.search(r'As', xlsx):
				ws.title = 'As Designed'
			elif re.search(r'FIXED', xlsx):
				ws.title = 'Fixed'
			elif re.search(r'CANCELLED',xlsx):
				ws.title = 'Cancelled'
			elif re.search(r'CLOSED', xlsx):
				ws.title = 'Closed'
			elif re.search(r'UNR', xlsx):
				ws.title = 'Unreproducable'

			workbook.add_sheet(ws)

		self.business_workbook = workbook


	def create_regular_workbook(self):

		workbook = Workbook()

		for xlsx in self.regular_xlsx:
			wb = openpyxl.load_workbook(xlsx)
			ws = wb.get_active_sheet()

			if re.search(r'Open\+bugs',xlsx):
				ws.title = 'Open'
			elif re.search(r'Resolved\+I', xlsx):
				ws.title = 'Resolved'
			elif re.search(r'Closed\+I', xlsx):
				ws.title = 'Closed'
			elif re.search(r'Open\+QC', xlsx):
				ws.title = 'Open QC'
			elif re.search(r'CANCELLED', xlsx):
				ws.title = 'Cancelled'
			elif re.search(r'Resolved\+QC', xlsx):
				ws.title = 'Resolved QC'
			elif re.search(r'Closed\+QC', xlsx):
				ws.title = 'Closed QC'

			workbook.add_sheet(ws)
		
		self.regular_workbook = workbook

	def clean_workbooks(self):

		sheet = self.business_workbook.get_sheet_by_name('Sheet')
		self.business_workbook.remove_sheet(sheet)

		sheet = self.regular_workbook.get_sheet_by_name('Sheet')
		self.regular_workbook.remove_sheet(sheet)


	def save_all_workbooks(self):
		self.business_workbook.save(os.path.join(self.current_directory, 'Delta_Athens_'+settings['date']+'_leadership.xlsx'))
		self.regular_workbook.save(os.path.join(self.current_directory, 'Delta_Athens_'+settings['date']+'.xlsx'))
